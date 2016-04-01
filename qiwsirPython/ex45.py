#!/usr/bin/env python
#coding:utf-8
#Created by Jeff on 2016/3/27 17:08

# 45、用python创建一个电子表格文件，并向其中填写数值。要求电子表格中有姓名、年龄、语言三项。实现对该电子表格文件数据的增加、删除、插入。

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "teacherInfo"

#add cell info
ws["A1"] = "Name"
ws["B1"] = "Age"
ws["C1"] = "Language"

for i in range(1,11):
    ws["A" + str(i+1)] = i
    ws["B" + str(i+1)] = i
    ws["C" + str(i+1)] = i

#delete rows data
for i in range(1,len(ws.rows[9])+1):
    ws.cell(row = 8 ,column = i).value = None
wb.save("example.xls")
